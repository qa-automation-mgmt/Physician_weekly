import pytest
import os
import time
import uuid
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from pytest_html import extras as html_extras

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

# ═══════════════════════════════════════════════════════
#  RUN FOLDER SETUP  (one timestamped folder per run)
# ═══════════════════════════════════════════════════════
RUN_TIME     = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
REPORT_DIR   = f"reports/Run_{RUN_TIME}"
SCREENSHOT_DIR = f"{REPORT_DIR}/screenshots"
HTML_REPORT  = f"{REPORT_DIR}/PW_Report.html"
EXCEL_FILE   = f"{REPORT_DIR}/PW_Execution_Report.xlsx"

os.makedirs(REPORT_DIR,    exist_ok=True)
os.makedirs(SCREENSHOT_DIR, exist_ok=True)

# ═══════════════════════════════════════════════════════
#  IN-MEMORY STORE  (populated by hook, consumed at end)
#
#  KEY DESIGN: dict keyed by test node-id so that every
#  rerun attempt OVERWRITES the previous entry.
#  Only the FINAL outcome (pass or fail) survives.
#  A test that fails twice then passes on retry → PASSED.
# ═══════════════════════════════════════════════════════
_test_results_map = {}   # nodeid -> result dict

# ═══════════════════════════════════════════════════════
#  PRIVACY POPUP HANDLER
# ═══════════════════════════════════════════════════════
def wait_and_close_privacy_popup(page, timeout=15000):
    close_btn = page.locator("#close-pc-btn-handler")
    try:
        close_btn.wait_for(state="visible", timeout=timeout)
        close_btn.click(force=True)
        page.wait_for_timeout(300)
        print("Privacy popup detected and closed")
    except PlaywrightTimeoutError:
        print("Privacy popup not shown")

# ═══════════════════════════════════════════════════════
#  BROWSER FIXTURE  (class-scoped – one browser per class)
# ═══════════════════════════════════════════════════════
@pytest.fixture(scope="class")
def page(request):
    playwright = sync_playwright().start()

    browser = playwright.chromium.launch(headless=True)

    context = browser.new_context(viewport={"width": 1920, "height": 1080})
    context.set_default_timeout(30000)
    context.set_default_navigation_timeout(30000)

    page = context.new_page()
    page.goto("https://www.physiciansweekly.com", wait_until="domcontentloaded")
    wait_and_close_privacy_popup(page)

    request.node.page = page
    yield page

    context.close()
    browser.close()
    playwright.stop()

# ═══════════════════════════════════════════════════════
#  HTML REPORT CUSTOMISATION
# ═══════════════════════════════════════════════════════
def pytest_html_report_title(report):
    report.title = "Physician Weekly – Automation Testing Report [Tester: Ashok]"

def pytest_html_results_summary(prefix, summary, postfix):
    prefix.extend([
        ("Project Name", "Physician Weekly"),
        ("Tester",       "Ashok Kumar"),
        ("URL",          "https://www.physiciansweekly.com/"),
        ("Browser",      "Chromium"),
        ("Environment",  "Production"),
    ])

# ═══════════════════════════════════════════════════════
#  PER-TEST HOOK  – collect results + screenshots
# ═══════════════════════════════════════════════════════
@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    report  = outcome.get_result()

    page = item.funcargs.get("page", None)

    if report.when == "call":
        status          = report.outcome.upper()   # PASSED / FAILED / ERROR
        failure_reason  = ""
        screenshot_path = ""
        manual_note     = ""
        remarks         = "Execution Completed Successfully"
        root_cause      = ""
        suggested_action = ""
        priority        = ""

        if report.failed:
            full_error = str(report.longrepr)

            # ── Failure classification ──────────────────────
            if "AssertionError" in full_error:
                failure_reason   = "Assertion validation failed"
                root_cause       = "Data / Logic Mismatch"
                suggested_action = "Verify element value / API response matches expected"
                priority         = "High"

            elif "TimeoutError" in full_error or "timeout" in full_error.lower():
                failure_reason   = "Element loading timeout or slow response"
                root_cause       = "Performance / Slow Load"
                suggested_action = "Add dynamic wait or increase timeout threshold"
                priority         = "Medium"

            elif "locator" in full_error.lower() and "strict mode" not in full_error.lower():
                failure_reason   = "Locator not found or locator changed"
                root_cause       = "Locator Change / UI Update"
                suggested_action = "Update XPath / CSS selector to match current DOM"
                priority         = "High"

            elif "strict mode violation" in full_error.lower():
                failure_reason   = "Multiple locators matched (strict mode)"
                root_cause       = "Ambiguous Locator"
                suggested_action = "Use more specific locator to target exactly one element"
                priority         = "Medium"

            elif "Element is not attached" in full_error:
                failure_reason   = "Element detached from DOM"
                root_cause       = "Stale Element Reference"
                suggested_action = "Re-query the element before interaction"
                priority         = "Medium"

            elif "Target page, context or browser has been closed" in full_error:
                failure_reason   = "Browser or page closed unexpectedly"
                root_cause       = "Environment / Stability Issue"
                suggested_action = "Check CI environment and browser resource limits"
                priority         = "High"

            elif "net::ERR" in full_error or "Navigation" in full_error:
                failure_reason   = "Navigation / Network error"
                root_cause       = "Network / URL Issue"
                suggested_action = "Verify URL is reachable and environment is stable"
                priority         = "High"

            else:
                failure_reason   = "Unexpected automation failure"
                root_cause       = "Unknown – Review full logs"
                suggested_action = "Review detailed error log and retry"
                priority         = "Low"

            manual_note = "✋ Verify manually before closing defect"
            remarks     = f"Failure – {failure_reason}"

            # ── Screenshot on failure ───────────────────────
            if page:
                screenshot_name = f"{item.name}_{int(time.time())}.png"
                screenshot_path = f"{SCREENSHOT_DIR}/{screenshot_name}"
                try:
                    page.screenshot(path=screenshot_path, full_page=True)
                except Exception as e:
                    print(f"Screenshot failed: {e}")
                    screenshot_path = ""

                # attach to HTML report
                if screenshot_path and not hasattr(report, "extras"):
                    report.extras = []
                if screenshot_path:
                    report.extras = getattr(report, "extras", [])
                    report.extras.append(html_extras.image(screenshot_path))

        # ── Store result (dict keyed by nodeid) ─────────────
        # Assigning by nodeid means any rerun attempt simply
        # OVERWRITES the previous failed entry for that test.
        # After all reruns are done, only the final outcome
        # (PASSED or truly-FAILED) survives in the dict.
        _test_results_map[item.nodeid] = {
            "Test Case Name":     item.name,
            "Module":             item.module.__name__.replace("test_", "").replace("_", " ").title(),
            "Status":             status,
            "Execution Time (s)": round(report.duration, 2),
            "Execution Date":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Browser":            "Chromium",
            "Environment":        "Production",
            "Failure Reason":     failure_reason,
            "Root Cause":         root_cause,
            "Suggested Action":   suggested_action,
            "Priority":           priority,
            "Screenshot Path":    screenshot_path,
            "Manual Verification": manual_note,
            "Remarks":            remarks,
        }

# ═══════════════════════════════════════════════════════
#  SESSION FINISH – GENERATE EXCEL REPORT
# ═══════════════════════════════════════════════════════
def pytest_sessionfinish(session, exitstatus):
    # Convert dict → list (only final outcomes, reruns already overwritten)
    final_results = list(_test_results_map.values())

    if not final_results:
        print("No test results collected – skipping Excel report.")
        return

    # ── Find the most recent previous Excel to carry run history forward ──
    previous_excel = None
    reports_root = "reports"
    if os.path.isdir(reports_root):
        all_runs = sorted([
            d for d in os.listdir(reports_root)
            if d.startswith("Run_") and os.path.isdir(os.path.join(reports_root, d))
        ])
        current_folder = os.path.basename(REPORT_DIR)
        for run_folder in reversed(all_runs):
            if run_folder == current_folder:
                continue
            candidate = os.path.join(reports_root, run_folder, "PW_Execution_Report.xlsx")
            if os.path.isfile(candidate):
                previous_excel = candidate
                break

    _generate_excel_report(final_results, EXCEL_FILE, RUN_TIME, previous_excel)
    print(f"\n{'='*60}")
    print(f"  Excel Report : {EXCEL_FILE}")
    print(f"  Screenshots  : {SCREENSHOT_DIR}")
    print(f"{'='*60}\n")


# ═══════════════════════════════════════════════════════
#  EXCEL REPORT BUILDER
# ═══════════════════════════════════════════════════════
def _generate_excel_report(results, filepath, run_time, previous_excel=None):

    df = pd.DataFrame(results)
    df.insert(0, "TC_ID", range(1, len(df) + 1))

    total    = len(df)
    passed   = int((df["Status"] == "PASSED").sum())
    failed   = int((df["Status"] == "FAILED").sum())
    errored  = int((df["Status"] == "ERROR").sum())
    pass_rate = round(passed / total * 100, 1) if total else 0
    total_time = round(df["Execution Time (s)"].sum(), 2)
    avg_time   = round(df["Execution Time (s)"].mean(), 2)

    # ── Colour palette ──────────────────────────────────
    C = {
        "pass_bg": "C6EFCE", "pass_fg": "276221",
        "fail_bg": "FFC7CE", "fail_fg": "9C0006",
        "warn_bg": "FFEB9C", "warn_fg": "9C5700",
        "title":   "1F3864", "header": "2E75B6",
        "subhdr":  "D6E4F0", "alt":    "F2F7FB",
        "white":   "FFFFFF", "border": "B8CCE4",
    }

    def fill(h):
        return PatternFill("solid", start_color=h, end_color=h)

    def fnt(bold=False, sz=9, color="000000"):
        return Font(bold=bold, size=sz, color=color, name="Arial")

    def bdr(style="thin", color="B8CCE4"):
        s = Side(style=style, color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def aln(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def hdr_row(ws, row, cols, bg="2E75B6", fg="FFFFFF", sz=9):
        for ci, val in enumerate(cols, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font   = Font(bold=True, size=sz, color=fg, name="Arial")
            c.fill   = fill(bg)
            c.border = bdr("thin", "FFFFFF")
            c.alignment = aln("center", "center", wrap=True)

    wb = Workbook()

    # ────────────────────────────────────────────────────
    #  SHEET 1 – EXECUTIVE SUMMARY
    # ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "📊 Executive Summary"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A6"

    # Title banner
    for r in range(1, 4):
        for cc in range(1, 13):
            ws1.cell(row=r, column=cc).fill = fill(C["title"])
    ws1.merge_cells("A1:L3")
    t = ws1["A1"]
    t.value     = "PHYSICIAN WEEKLY – AUTOMATION EXECUTION REPORT"
    t.font      = Font(bold=True, size=20, color="FFFFFF", name="Arial")
    t.alignment = aln("center", "center")

    ws1.merge_cells("A4:L4")
    s = ws1["A4"]
    s.value = (
        f"Project: Physician Weekly  |  Tester: Ashok Kumar  |  "
        f"Run: {run_time}  |  Browser: Chromium  |  Environment: Production"
    )
    s.font      = Font(italic=True, size=10, color="FFFFFF", name="Arial")
    s.fill      = fill(C["header"])
    s.alignment = aln("center", "center")

    for r in [1,2,3,4]:
        ws1.row_dimensions[r].height = 16

    # KPI Cards (row 6 label, rows 7-10 value)
    ws1.cell(row=5, column=1).value = "EXECUTION METRICS"
    ws1.cell(row=5, column=1).font  = fnt(bold=True, sz=11, color=C["header"])

    kpis = [
        ("Total Test Cases", total,            "BDD7EE", "000000"),
        ("✅  Passed",       passed,            C["pass_bg"], C["pass_fg"]),
        ("❌  Failed",       failed,            C["fail_bg"], C["fail_fg"]),
        ("📈  Pass Rate",    f"{pass_rate}%",   "FFEB9C" if pass_rate < 100 else C["pass_bg"],
                                               C["warn_fg"] if pass_rate < 100 else C["pass_fg"]),
        ("⏱  Total Time",   f"{total_time}s",  "BDD7EE", "000000"),
        ("⚡  Avg Time",     f"{avg_time}s",    "BDD7EE", "000000"),
    ]
    col_pairs = [1, 3, 5, 7, 9, 11]
    for idx, (label, value, bg, fg) in enumerate(kpis):
        sc = col_pairs[idx]
        lc = ws1.cell(row=6, column=sc, value=label)
        lc.font      = Font(bold=True, size=9, color="FFFFFF", name="Arial")
        lc.fill      = fill(C["header"])
        lc.alignment = aln("center", "center")
        lc.border    = bdr("medium", "1F3864")
        ws1.merge_cells(start_row=6, start_column=sc, end_row=6, end_column=sc+1)

        vc = ws1.cell(row=7, column=sc, value=value)
        vc.font      = Font(bold=True, size=22, color=fg, name="Arial")
        vc.fill      = fill(bg)
        vc.alignment = aln("center", "center")
        vc.border    = bdr("medium", "1F3864")
        ws1.merge_cells(start_row=7, start_column=sc, end_row=10, end_column=sc+1)

    for r in range(6, 11):
        ws1.row_dimensions[r].height = 16

    # ── Failed test detail ──────────────────────────────
    ws1.cell(row=12, column=1).value = "⚠️  FAILED TEST CASES"
    ws1.cell(row=12, column=1).font  = fnt(bold=True, sz=12, color=C["fail_fg"])
    ws1.merge_cells("A12:L12")

    hdr_row(ws1, 13, ["TC_ID","Test Case Name","Module","Failure Reason","Root Cause","Suggested Action","Priority","Screenshot"], bg=C["fail_fg"])

    failed_df = df[df["Status"] == "FAILED"]
    r_idx = 14
    for _, row in failed_df.iterrows():
        prio = row["Priority"]
        prio_bg = C["fail_bg"] if prio == "High" else (C["warn_bg"] if prio == "Medium" else "BDD7EE")
        vals = [
            row["TC_ID"], row["Test Case Name"], row["Module"],
            row["Failure Reason"] or "–",
            row["Root Cause"]      or "–",
            row["Suggested Action"] or "–",
            prio,
            "📸 Captured" if row["Screenshot Path"] else "–",
        ]
        for ci, val in enumerate(vals, 1):
            c = ws1.cell(row=r_idx, column=ci, value=val)
            c.font      = fnt(sz=9, color=C["fail_fg"])
            c.fill      = fill(prio_bg if ci == 7 else C["fail_bg"])
            c.border    = bdr()
            c.alignment = aln("left" if ci in [2,4,5,6] else "center", "center", wrap=True)
        ws1.row_dimensions[r_idx].height = 28
        r_idx += 1

    if len(failed_df) == 0:
        c = ws1.cell(row=14, column=1, value="🎉  All test cases PASSED! No failures to report.")
        c.font  = Font(bold=True, size=11, color=C["pass_fg"], name="Arial")
        c.fill  = fill(C["pass_bg"])
        ws1.merge_cells("A14:L14")
        r_idx = 15

    # ── Module breakdown ────────────────────────────────
    mstart = r_idx + 2
    ws1.cell(row=mstart, column=1).value = "MODULE-WISE BREAKDOWN"
    ws1.cell(row=mstart, column=1).font  = fnt(bold=True, sz=12, color=C["header"])

    hdr_row(ws1, mstart+1, ["Module","Total","Passed","Failed","Pass Rate","Result"])
    mod_grp = df.groupby("Module").agg(
        Total=("Status","count"),
        Passed=("Status", lambda x: (x=="PASSED").sum()),
        Failed=("Status", lambda x: (x=="FAILED").sum()),
    ).reset_index()
    mod_grp["Pass Rate"] = (mod_grp["Passed"] / mod_grp["Total"] * 100).round(1)

    for ri, (_, mr) in enumerate(mod_grp.iterrows(), mstart+2):
        ok   = mr["Failed"] == 0
        is_alt = ri % 2 == 0
        vals = [mr["Module"], mr["Total"], mr["Passed"], mr["Failed"],
                f"{mr['Pass Rate']}%", "✅ PASS" if ok else "❌ FAIL"]
        bgs  = [C["alt"] if is_alt else C["white"],
                C["white"], C["pass_bg"] if mr["Passed"] else C["white"],
                C["fail_bg"] if mr["Failed"] else C["white"],
                C["white"], C["pass_bg"] if ok else C["fail_bg"]]
        fgs  = ["000000","000000",
                C["pass_fg"] if mr["Passed"] else "000000",
                C["fail_fg"] if mr["Failed"] else "000000",
                "000000", C["pass_fg"] if ok else C["fail_fg"]]
        for ci, (v, b, f) in enumerate(zip(vals, bgs, fgs), 1):
            c = ws1.cell(row=ri, column=ci, value=v)
            c.fill      = fill(b)
            c.font      = Font(bold=(ci==6), size=9, color=f, name="Arial")
            c.border    = bdr()
            c.alignment = aln("left" if ci==1 else "center","center")
        ws1.row_dimensions[ri].height = 18

    # ── Pie chart ───────────────────────────────────────
    chart_data_row = mstart + len(mod_grp) + 5
    ws1.cell(row=chart_data_row,   column=1, value="Status")
    ws1.cell(row=chart_data_row,   column=2, value="Count")
    ws1.cell(row=chart_data_row+1, column=1, value="Passed")
    ws1.cell(row=chart_data_row+1, column=2, value=passed)
    ws1.cell(row=chart_data_row+2, column=1, value="Failed")
    ws1.cell(row=chart_data_row+2, column=2, value=failed if failed else 0)

    pie = PieChart()
    pie.title  = "Pass vs Fail"
    pie.style  = 10
    pie.width  = 14
    pie.height = 10
    d_ref = Reference(ws1, min_col=2, min_row=chart_data_row,   max_row=chart_data_row+2)
    l_ref = Reference(ws1, min_col=1, min_row=chart_data_row+1, max_row=chart_data_row+2)
    pie.add_data(d_ref, titles_from_data=True)
    pie.set_categories(l_ref)
    dp_pass = DataPoint(idx=0); dp_pass.graphicalProperties.solidFill = "70AD47"
    dp_fail = DataPoint(idx=1); dp_fail.graphicalProperties.solidFill = "FF0000"
    pie.series[0].dPt = [dp_pass, dp_fail]
    ws1.add_chart(pie, "H6")

    # ── Bar chart (execution time) ──────────────────────
    bc_start = chart_data_row + 5
    ws1.cell(row=bc_start, column=1, value="Test #")
    ws1.cell(row=bc_start, column=2, value="Time (s)")
    for i, (_, row) in enumerate(df.iterrows()):
        ws1.cell(row=bc_start+1+i, column=1, value=int(row["TC_ID"]))
        ws1.cell(row=bc_start+1+i, column=2, value=row["Execution Time (s)"])

    bar = BarChart()
    bar.type    = "bar"
    bar.title   = "Execution Time per Test (s)"
    bar.style   = 10
    bar.width   = max(28, len(df) * 0.4)   # scale width with test count
    bar.height  = max(14, len(df) * 0.22)  # scale height so bars don't squash
    bar.y_axis.title = "Seconds"
    bar.x_axis.title = "TC_ID"
    # Force every TC_ID label to show — no skipping
    bar.x_axis.tickLblSkip  = 1
    bar.x_axis.tickMarkSkip = 1
    bar.x_axis.noMultiLvlLbl = True
    td = Reference(ws1, min_col=2, min_row=bc_start, max_row=bc_start+len(df))
    tc = Reference(ws1, min_col=1, min_row=bc_start+1, max_row=bc_start+len(df))
    bar.add_data(td, titles_from_data=True)
    bar.set_categories(tc)
    ws1.add_chart(bar, "A" + str(mstart + len(mod_grp) + 4))

    # column widths for sheet 1
    for ci, w in enumerate([6,10,14,14,14,14,12,12], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ────────────────────────────────────────────────────
    #  SHEET 2 – DETAILED EXECUTION LOG
    # ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("📋 Execution Log")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A4"

    ws2.merge_cells("A1:L2")
    t2 = ws2["A1"]
    t2.value     = f"DETAILED EXECUTION LOG – Physician Weekly  |  Run: {run_time}"
    t2.font      = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    t2.fill      = fill(C["title"])
    t2.alignment = aln("center", "center")
    for r in [1,2]:
        ws2.row_dimensions[r].height = 16

    log_hdrs = ["TC_ID","Test Case Name","Module","Status","Time (s)",
                "Execution Date","Browser","Environment","Failure Reason","Remarks","Screenshot"]
    hdr_row(ws2, 3, log_hdrs)
    ws2.row_dimensions[3].height = 30

    log_widths = [6, 45, 22, 10, 10, 22, 10, 12, 40, 35, 30]
    for ci, w in enumerate(log_widths, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    for r_i, (_, row) in enumerate(df.iterrows(), 4):
        is_pass = row["Status"] == "PASSED"
        alt     = r_i % 2 == 0
        base    = C["alt"] if alt else C["white"]
        vals = [
            row["TC_ID"], row["Test Case Name"], row["Module"],
            "✅ PASS" if is_pass else "❌ FAIL",
            row["Execution Time (s)"], row["Execution Date"],
            row["Browser"], row["Environment"],
            row["Failure Reason"] or "–", row["Remarks"],
        ]
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=r_i, column=ci, value=val)
            c.border    = bdr()
            c.alignment = aln("left" if ci in [2,3,9,10] else "center","center", wrap=(ci in [9,10]))

            if ci == 4:                         # Status
                c.fill = fill(C["pass_bg"] if is_pass else C["fail_bg"])
                c.font = Font(bold=True, size=9, color=C["pass_fg"] if is_pass else C["fail_fg"], name="Arial")
            elif ci == 5:                       # Time – green/amber/red
                t_val = row["Execution Time (s)"]
                t_bg  = C["pass_bg"] if t_val < 10 else (C["warn_bg"] if t_val < 30 else C["fail_bg"])
                c.fill = fill(t_bg)
                c.font = fnt(sz=9)
            elif ci == 9 and not is_pass:       # Failure reason
                c.fill = fill(C["warn_bg"])
                c.font = fnt(sz=9, color=C["warn_fg"])
            else:
                c.fill = fill(base)
                c.font = fnt(sz=9)
        # ── Column 11: Screenshot relative path (failures only) ──
        # Stored as a relative path so the link works on ANY machine
        # as long as the recipient has the full run folder (Excel + screenshots/).
        # Relative path is: screenshots/test_name_timestamp.png
        sc = ws2.cell(row=r_i, column=11)
        sc.border    = bdr()
        sc.alignment = aln("center", "center")
        shot = row.get("Screenshot Path", "")
        if not is_pass and shot:
            # Make path relative to the report folder so it works on any machine
            rel_path = os.path.relpath(shot, start=os.path.dirname(filepath))
            # Use forward slashes — works on Windows, Linux, Mac
            rel_path = rel_path.replace("\\", "/")
            sc.value     = rel_path
            sc.hyperlink = rel_path
            sc.font      = Font(size=9, color="0563C1", underline="single", name="Arial")
            sc.fill      = fill(C["fail_bg"])
        else:
            sc.value = "–"
            sc.font  = fnt(sz=9, color="999999")
            sc.fill  = fill(base)

        ws2.row_dimensions[r_i].height = 20

    # Totals footer
    fr = 4 + len(df)
    ws2.cell(row=fr, column=1, value="TOTALS →").font = fnt(bold=True, sz=9)
    ws2.cell(row=fr, column=4, value=f"P:{passed}  F:{failed}").font = fnt(bold=True, sz=9, color=C["header"])
    ws2.cell(row=fr, column=5, value=f"{total_time}s").font = fnt(bold=True, sz=9)
    for ci in range(1, 12):
        ws2.cell(row=fr, column=ci).fill   = fill(C["subhdr"])
        ws2.cell(row=fr, column=ci).border = bdr("medium", C["header"])

    # ────────────────────────────────────────────────────
    #  SHEET 3 – FAILURE ANALYSIS
    # ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("🔍 Failure Analysis")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:H2")
    t3 = ws3["A1"]
    t3.value     = "FAILURE ANALYSIS & DEFECT TRIAGE – Physician Weekly"
    t3.font      = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    t3.fill      = fill("C00000")
    t3.alignment = aln("center", "center")
    for r in [1,2]:
        ws3.row_dimensions[r].height = 16

    fa_hdrs = ["TC_ID","Test Case Name","Module","Failure Reason",
               "Root Cause","Suggested Action","Priority","Defect Status"]
    hdr_row(ws3, 3, fa_hdrs, bg="C00000")

    fa_widths = [6, 42, 22, 45, 25, 45, 10, 14]
    for ci, w in enumerate(fa_widths, 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    r_idx = 4
    for _, row in failed_df.iterrows():
        prio    = row["Priority"]
        prio_bg = C["fail_bg"] if prio=="High" else (C["warn_bg"] if prio=="Medium" else "BDD7EE")
        prio_fg = C["fail_fg"] if prio=="High" else (C["warn_fg"] if prio=="Medium" else "000000")
        vals = [
            row["TC_ID"], row["Test Case Name"], row["Module"],
            row["Failure Reason"] or "–",
            row["Root Cause"]      or "–",
            row["Suggested Action"] or "–",
            prio, "🔴 Open",
        ]
        for ci, val in enumerate(vals, 1):
            c = ws3.cell(row=r_idx, column=ci, value=val)
            c.border    = bdr()
            c.alignment = aln("left" if ci in [2,4,5,6] else "center","center",wrap=True)
            if ci == 7:
                c.fill = fill(prio_bg); c.font = Font(bold=True,size=9,color=prio_fg,name="Arial")
            elif ci == 8:
                c.fill = fill(C["fail_bg"]); c.font = fnt(bold=True,sz=9,color=C["fail_fg"])
            else:
                c.fill = fill("FFF9F9"); c.font = fnt(sz=9)
        ws3.row_dimensions[r_idx].height = 30
        r_idx += 1

    if len(failed_df) == 0:
        c = ws3.cell(row=4, column=1, value="✅  No failures in this run. Great job!")
        c.font  = Font(bold=True, size=12, color=C["pass_fg"], name="Arial")
        c.fill  = fill(C["pass_bg"])
        ws3.merge_cells("A4:H4")

    # ────────────────────────────────────────────────────
    #  SHEET 4 – RUN HISTORY  (auto-accumulated across runs)
    # ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("📝 Run History")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:G2")
    t4 = ws4["A1"]
    t4.value     = "EXECUTION RUN HISTORY – Physician Weekly (Auto-accumulated)"
    t4.font      = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    t4.fill      = fill(C["title"])
    t4.alignment = aln("center", "center")
    for r in [1,2]:
        ws4.row_dimensions[r].height = 16

    rh_hdrs = ["Run #","Run Timestamp","Total TCs","Passed","Failed","Pass Rate (%)","Notes / Changes"]
    hdr_row(ws4, 3, rh_hdrs)
    rh_widths = [8, 22, 10, 10, 10, 14, 55]
    for ci, w in enumerate(rh_widths, 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    # ── Load previous run history rows from the last Excel ──
    previous_rows = []
    if previous_excel:
        try:
            from openpyxl import load_workbook as _lw
            _prev_wb = _lw(previous_excel, read_only=True, data_only=True)
            _prev_ws = None
            for _sname in _prev_wb.sheetnames:
                if "Run History" in _sname:
                    _prev_ws = _prev_wb[_sname]
                    break
            if _prev_ws:
                for _row in _prev_ws.iter_rows(min_row=4, values_only=True):
                    # keep only rows where col A is an integer (the Run # column)
                    try:
                        int(_row[0])
                        previous_rows.append(list(_row[:7]))  # only 7 history columns
                    except (TypeError, ValueError):
                        pass  # skip tip rows, blank rows, merged cell artefacts
            _prev_wb.close()
        except Exception as e:
            print(f"Warning: could not read previous run history – {e}")

    # ── Write all previous rows first, then append current run ──
    next_run_num = len(previous_rows) + 1
    all_history_rows = previous_rows + [
        [next_run_num, run_time, total, passed, failed, f"{pass_rate}%",
         f"Auto-generated | Failures: {failed} | Avg time: {avg_time}s"]
    ]

    for row_offset, hist_row in enumerate(all_history_rows):
        excel_row = 4 + row_offset
        is_current = (row_offset == len(all_history_rows) - 1)
        row_bg = "FFFDE7" if is_current else (C["alt"] if row_offset % 2 == 0 else C["white"])
        for ci, val in enumerate(hist_row, 1):
            c = ws4.cell(row=excel_row, column=ci, value=val)
            c.font      = fnt(sz=9, bold=is_current)
            c.border    = bdr()
            c.fill      = fill(row_bg)
            c.alignment = aln("center","center")
        ws4.row_dimensions[excel_row].height = 18

    # blank buffer rows after history
    last_data_row = 4 + len(all_history_rows)
    for r in range(last_data_row, last_data_row + 10):
        for ci in range(1, 8):
            c = ws4.cell(row=r, column=ci)
            c.border = bdr("hair","CCCCCC")
            c.fill   = fill("FAFAFA")
        ws4.row_dimensions[r].height = 18

    tip_row = last_data_row + 11
    tip = ws4.cell(row=tip_row, column=1,
                   value="💡 Run history is auto-accumulated. Each new run appends automatically.")
    tip.font = Font(italic=True, size=9, color="666666", name="Arial")
    ws4.merge_cells(f"A{tip_row}:G{tip_row}")

    wb.save(filepath)